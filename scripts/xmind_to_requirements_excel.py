#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse XMind content.json and export 端口-板块-功能点-详细功能描述 to xlsx."""

from __future__ import annotations

import json
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def get_title(topic: dict) -> str:
    t = topic.get("title", "")
    if isinstance(t, str):
        return t.strip()
    return str(t).strip()


def get_notes(topic: dict) -> str:
    n = topic.get("notes")
    if not n:
        return ""
    if isinstance(n, dict):
        for key in ("plain", "realHTML", "html"):
            v = n.get(key)
            if isinstance(v, str) and v.strip():
                return v.strip()
            if isinstance(v, dict):
                c = v.get("content")
                if isinstance(c, str) and c.strip():
                    return c.strip()
    return str(n).strip()


def children(topic: dict) -> list[dict]:
    ch = topic.get("children") or {}
    return ch.get("attached") or []


def join_items(items: list[str], limit: int = 18) -> str:
    shown = [" ".join(item.split()) for item in items[:limit]]
    text = "、".join(shown)
    if len(items) > limit:
        text += f"等共{len(items)}项"
    return text


def infer_item_type(title: str, child_titles: list[str]) -> str:
    text = title + " " + " ".join(child_titles)
    if any(k in text for k in ("登录", "注册", "认证", "授权", "密码")):
        return "账号与认证"
    if any(k in text for k in ("搜索", "筛选", "排序", "查询", "检索")):
        return "查询筛选"
    if any(k in text for k in ("列表", "记录", "明细", "订单", "用户列表", "商品列表")):
        return "列表展示"
    if any(k in text for k in ("操作按钮", "添加", "删除", "修改", "编辑", "冻结", "解封", "刷新", "提交", "确认")):
        return "操作管理"
    if any(k in text for k in ("状态", "分类", "待支付", "已完成", "退款", "待使用", "待评价")):
        return "状态分类"
    if any(k in text for k in ("金额", "价格", "支付", "充值", "余额", "积分", "财务", "营收")):
        return "交易财务"
    if any(k in text for k in ("Banner", "广告", "内容", "图文", "介绍", "海报")):
        return "内容展示"
    if any(k in text for k in ("房间", "房型", "入住", "退房", "打扫", "保洁")):
        return "酒店业务"
    if child_titles:
        return "页面模块"
    return "字段信息"


def build_description(port: str, plate: str, point: str, path: list[str], child_titles: list[str], notes: str) -> str:
    parent = path[-2] if len(path) >= 2 else ""
    item_type = infer_item_type(point, child_titles)
    endpoint = port if port.endswith("端") or port.endswith("工作台") or port in ("开发总览", "管理端") else f"{port}端"
    location = endpoint
    if plate:
        location += f"的「{plate}」板块"
    else:
        location += "入口"

    desc: list[str] = []
    if child_titles:
        desc.append(
            f"在{location}中提供「{point}」{item_type}，用于承载与「{point}」相关的页面展示、信息录入、业务处理和操作入口。"
        )
        desc.append(f"页面/模块应包含：{join_items(child_titles)}。")
        if item_type == "账号与认证":
            desc.append("业务逻辑：支持用户或员工按导图所列方式完成身份识别、信息校验和登录/绑定流程；未满足必填或授权条件时应给出明确提示。")
        elif item_type == "查询筛选":
            desc.append("业务逻辑：支持按关键字、位置、时间、状态或导图列出的筛选项检索数据，并将查询结果联动到对应列表或详情页。")
        elif item_type == "列表展示":
            desc.append("业务逻辑：以列表方式展示核心数据，每条记录按导图字段呈现关键信息，并根据业务状态提供可用操作。")
        elif item_type == "操作管理":
            desc.append("业务逻辑：提供新增、编辑、删除、提交、刷新或状态变更等操作入口；操作后需要更新当前页面数据并反馈处理结果。")
        elif item_type == "状态分类":
            desc.append("业务逻辑：按导图列出的状态或分类对数据进行分组展示，用户可切换分类查看对应数据范围。")
        elif item_type == "交易财务":
            desc.append("业务逻辑：围绕金额、支付、积分、余额或报表数据完成展示、选择、计算和结果确认，涉及支付或财务数据时需保证金额清晰可核对。")
        elif item_type == "内容展示":
            desc.append("业务逻辑：用于展示平台内容、活动、广告或图文信息，支持按配置内容在前端页面呈现。")
        elif item_type == "酒店业务":
            desc.append("业务逻辑：围绕酒店预订、入住、退房、房间、保洁或服务场景完成信息展示、流程流转和业务操作。")
        else:
            desc.append("业务逻辑：根据导图下级节点组织页面结构，用户可查看、填写或操作相关信息，系统按模块关系完成信息展示与流程跳转。")
    else:
        desc.append(f"「{point}」作为{location}中「{parent}」的具体字段、信息项或操作项使用。")
        desc.append("页面逻辑：在对应页面中用于展示、录入、选择或触发该项内容，需与所在父级功能保持联动。")

    if notes:
        desc.append(f"补充说明：{notes}")
    desc.append(f"来源路径：{' > '.join(path)}")
    return "\n".join(desc)


def walk(topic: dict, path: list[str], rows: list[dict]) -> None:
    title = get_title(topic)
    if not title:
        return
    cur_path = path + [title]
    # 跳过仅「根 > 端口」的节点，避免出现 端口=用户端 且 功能点=用户端 的重复行
    if len(cur_path) >= 3:
        port = cur_path[1]
        if len(cur_path) >= 4:
            plate = " / ".join(cur_path[2:-1])
        else:
            plate = ""
        point = cur_path[-1]
        notes = get_notes(topic)
        subs = [get_title(c) for c in children(topic)]
        subs = [s for s in subs if s]
        description = build_description(port, plate, point, cur_path, subs, notes)
        rows.append(
            {
                "端口": port,
                "板块": plate,
                "功能点": point,
                "详细功能描述": description,
            }
        )
    for c in children(topic):
        walk(c, cur_path, rows)


def load_root_from_xmind(xmind: Path) -> dict:
    with zipfile.ZipFile(xmind, "r") as zf:
        names = zf.namelist()
        if "content.json" in names:
            data = json.loads(zf.read("content.json").decode("utf-8"))
            if isinstance(data, list) and data:
                return data[0]["rootTopic"]
            if isinstance(data, dict) and "rootTopic" in data:
                return data["rootTopic"]
        if "content.xml" in names:
            raise SystemExit("此文件为 XML 格式，请先用 XMind 另存为含 content.json 的版本。")
    raise SystemExit("未找到 content.json")


def export_xlsx(rows: list[dict], out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "需求功能"
    headers = ["端口", "板块", "功能点", "详细功能描述"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r["端口"], r["板块"], r["功能点"], r["详细功能描述"]])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [14, 36, 28, 72]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(out)


def main() -> None:
    if len(sys.argv) < 3:
        print("用法: xmind_to_requirements_excel.py <输入.xmind> <输出.xlsx>", file=sys.stderr)
        sys.exit(2)
    src = Path(sys.argv[1]).expanduser()
    dst = Path(sys.argv[2]).expanduser()
    root = load_root_from_xmind(src)
    rows: list[dict] = []
    walk(root, [], rows)
    export_xlsx(rows, dst)
    print(f"已写入 {dst} ，数据行 {len(rows)} 条（不含表头）")


if __name__ == "__main__":
    main()
