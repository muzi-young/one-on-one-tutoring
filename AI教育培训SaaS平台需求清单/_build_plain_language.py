#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从主需求 xlsx 生成通俗版 xlsx/csv（六列含备注；备注源表无则留空）。"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
SRC = BASE / "AI教育培训SaaS平台-需求功能清单.xlsx"
OUT_XLSX = BASE / "AI教育培训SaaS平台-需求功能清单_通俗版.xlsx"
OUT_CSV = BASE / "AI教育培训SaaS平台-需求功能清单_通俗版.csv"
OUT_TXT = BASE / "通俗版_使用说明.txt"

# ---------- 词表与替换（顺序敏感：长串优先）----------
ORDERED_REPLACEMENTS: list[tuple[str, str]] = [
    (r"SendGrid/SES", "SendGrid 或云服务商邮箱接口"),
    (r"\bSES\b", "云平台邮件收发（亚马逊 SES 类）"),
    (r"APNs/FCM", "苹果推送与安卓推送"),
    (r"APNs", "苹果推送通知"),
    (r"\bFCM\b", "安卓推送通知"),
    (r"JWT \+ Refresh Token", "登录令牌与续期令牌（保持登录态用）"),
    (r"\bJWT\b", "登录令牌（短时有效）"),
    (r"Asia/Tashkent", "塔什干时间（乌兹别克）"),
    (r"Tashkent", "塔什干"),
    (r"Payme/Click/Uzum/Humo/Uzcard/Apelsin", "Payme / Click 等本地支付（含 Uzum / Humo / Uzcard / Apelsin，按阶段接入）"),
    (r"Payme/Click", "Payme 或 Click 支付"),
    (r"Humo/Uzcard", "Humo 或银行卡"),
    (r"\bCloudFront\b", "亚马逊内容分发网络 CloudFront"),
    (r"Bunny CDN", "Bunny 内容分发加速"),
    (r"CloudFront/Bunny CDN", "内容分发加速（中亚体验优先时可选 Bunny 等）"),
    (r"\bCDN\b", "内容分发加速"),
    (r"\bAWS\b", "亚马逊云"),
    (r"AWS S3", "云对象存储"),
    (r"S3 ", "对象存储 "),
    (r"→ S3", "→ 对象存储"),
    (r"\bKMS\b", "云端密钥保险箱"),
    (r"\bAK/SK\b", "云平台访问密钥"),
    (r"\beu-central-1\b", "法兰克福区域"),
    (r"\bap-south-1\b", "孟买区域"),
    (r"MediaConvert", "云端视频转码"),
    (r"\bSDK\b", "音视频互动组件"),
    (r"\bAPI\b", "标准程序接口"),
    (r"Merchant API", "商户支付接口"),
    (r"\bWebhook\b", "外部回调通知地址"),
    (r"\bUTC\+5\b", "塔什干时区（较世界协调时早 5 小时）"),
    (r"\bUTC\b", "世界协调时"),
    (r"\bISO 8601\b", "国际通用时间字符串格式"),
    (r"yyyy-MM-dd HH:mm", "年-月-日 时:分"),
    (r"ISO 8601 内部存储 UTC", "系统内部用世界协调时统一存时间"),
    (r"\bRAG\b", "先检索知识片段再让大模型作答"),
    (r"\bLLM\b", "大语言模型"),
    (r"\bVTT\b", "网页字幕文件"),
    (r"\bOCR\b", "图片文字识别"),
    (r"\bURL\b", "链接地址"),
    (r"\bPNG\b", "图片"),
    (r"\bPDF\b", "PDF 文档"),
    (r"\bics\b", "日历订阅文件"),
    (r"\bIAP\b", "应用内购买"),
    (r"\bSKU\b", "商品款号"),
    (r"\bROI\b", "投入产出"),
    (r"\bKPI\b", "关键指标"),
    (r"\bOKR\b", "目标与关键结果"),
    (r"\bSLA\b", "服务承诺时限"),
    (r"\bA/B\b", "对照试验"),
    (r"\bFAQ\b", "常见问题"),
    (r"\bOAuth\b", "第三方授权登录协议"),
    (r"\bCRM\b", "客户管理"),
    (r"\bICS\b", "日历同步文件"),
    (r"iOS 14\+\s*ATT\s*弹窗合规", "苹果 iOS14 及以上：首次须按系统弹出「是否允许广告追踪」询问（行业常用缩写 ATT）"),
    (r"\bATT \(App Tracking Transparency\)", "苹果「应用追踪透明度」提示"),
    # 须在泛化 ATT 之前处理授权连用
    (r"\bATT\s*授权", "苹果广告追踪授权"),
    (r"Privacy Manifest", "苹果隐私清单"),
    (r"PrivacyInfo\.xcprivacy", "苹果隐私配置文件"),
    (r"GDPR", "欧盟个人信息保护通用规则"),
    (r"COPPA", "美国儿童上网隐私相关规则"),
    (r"ZRU-547", "乌兹别克个人信息/数据权利相关法律"),
    (r"Consent Banner", "Cookies 与统计营销同意条"),
    (r"DKIM/SPF/DMARC", "邮箱防伪与域名可信三项配置"),
    (r"Mixpanel / Amplitude / GA4", "产品分析工具（任选其一，按选型）"),
    (r"Pinecone/Weaviate/Qdrant/pgvector", "向量库存储（按项目实际选型）"),
    (r"OpenAI/Claude/Gemini", "常用大模型云服务（按合规与选型确定）"),
    (r"Branch\.io / Firebase Dynamic Links / AppsFlyer OneLink", "第三方短链与安装归因工具（Branch / Firebase 动态链接 / AppsFlyer OneLink，按选型）"),
    (r"Firebase Dynamic Links", "Firebase 安装短链"),
    (r"\bRTC\b", "实时音视频链路"),
    (r"\bJSON\b", "JSON 数据包（常用交换格式）"),
    (r"\bUGC\b", "学员原创上传内容"),
    (r"\bBI\b", "商业智能统计分析"),
    (r"\bPPT\b", "演示文稿"),
    (r"\bSMS\b", "短信"),
    (r"Agora / 100ms / LiveKit", "实时音视频服务（如声网等，按选型）"),
    (r"Branch", "第三方深度链接与归因服务 Branch"),
    (r"Intercom 或 Freshdesk", "在线客服系统（Intercom 或 Freshdesk 等）"),
    (r"Whisper", "语音转写模型 Whisper"),
    (r"DeepL/Google", "机器翻译服务（DeepL / 谷歌等）"),
    (r"\bDocuSign\b", "DocuSign 电子签"),
    (r"\be-IMZO\b", "乌兹别克电子签章方式 e-IMZO"),
    (r"\bT\+1\b", "成交后次日"),
    (r"\bT\+N\b", "成交后第 N 个工作日"),
    (r"\bUZS\b", "乌兹别克苏姆"),
    (r"QQS 12%", "乌兹别克增值税约 12%（QQS，税率以法务为准）"),
    (r"QQS 增值税", "乌兹别克增值税（QQS）"),
    (r"乌兹别克 QQS", "乌兹别克增值税（当地惯称 QQS）"),
    (r"\(QQS 12%\)", "（增值税约 12%，QQS）"),
    (r"发票管理\(QQS 12%\)", "发票管理（乌兹别克增值税 QQS）"),
    (r"\bQQS\b", "增值税（QQS）"),
    (r"永久绑定写入", "永久绑定关系登记"),
    (r"\bK12\b", "中小学（小学至高中）"),
    (r"Device ID\(IDFA/AAID\)", "设备标识（苹果 IDFA / 谷歌广告 ID）"),
    (r"IDFA/AAID", "苹果广告标识 / 安卓广告标识"),
    (r"\bIDFA\b", "苹果广告标识"),
    (r"\bAAID\b", "安卓广告标识"),
    (r"chat_id", "对话编号"),
    (r"Telegram 对话编号 绑定", "绑定 Telegram（用于识别会话）"),
    (r"/verify", "在机器人里发送验证指令"),
    (r"Verifiable", "可核验"),
    (r"\bCIS\b", "独联体地区"),
    (r"\b360p/480p/720p/1080p\b", "多档清晰度（360p 至 1080p）"),
    (r"0\.75/1/1\.25/1\.5/2", "0.75～2 倍速"),
    (r"\bEmail\b", "邮箱"),
    (r"\bemail\b", "邮箱"),
    (r"Telegram Bot Token", "Telegram 机器人密钥"),
    (r"\bBot Token\b", "机器人密钥"),
    (r"Telegram Bot", "Telegram 机器人"),
    (r"复用学员侧 在机器人里发送验证指令 流程", "复用学员侧 Telegram 机器人验证流程"),
    (r"\bTG\b(?![a-z])", "Telegram"),
    (r"\bTelegram\b", "Telegram"),
    (r"\bApp Store\b", "苹果应用商店"),
    (r"\bStore\b", "应用商店"),
    (r"App 审核", "应用上架审核"),
    (r"ru/uz/en/zh-CN", "俄/乌/英/中"),
    (r"ru/uz/en/zh", "俄/乌/英/中"),
    (r"ru/uz", "俄/乌"),
]

# 开发向口语 → 客户向
DEV_PHRASES: list[tuple[str, str]] = [
    (r"登录成功下发", "登录成功后，系统向对方发放"),
    (r"校验", "检查"),
    (r"写入", "保存"),
    (r"落库", "写入系统存档"),
    (r"下发", "发送到"),
    (r"拦截", "阻止并提示"),
    (r"命中", "匹配到"),
    (r"召回", "搜出相关条目"),
    (r"重排", "再排序"),
    (r"渲染", "生成页面或文件"),
    (r"留痕", "记录备查"),
    (r"结构化入库", "结构化后写入系统"),
    (r"冲减", "扣回"),
    (r"回写", "同步更新显示"),
    (r"聚合", "汇总"),
    (r"扫描", "定期查找"),
    (r"限频", "限制频率"),
    (r"风控", "风险监控"),
    (r"会话", "登录会话"),
    (r"会话 token", "登录凭证"),
    (r"\btoken\b", "临时口令"),
    (r"Token", "密钥或口令"),
    (r"uid", "用户编号"),
    (r"强制下线所有会话", "全部设备重新登录"),
]

# 功能点/板块里英文括号弱化
POINT_EXTRA: list[tuple[str, str]] = [
    (r"\(Telegram 唯一\)", "（以 Telegram 为主通道）"),
    (r"\(学生端\)", ""),
    (r"\(学员侧\)", ""),
    (r"\(教师侧\)", ""),
    (r"\(家长端\)", ""),
]


def _sub_chain(text: str, pairs: list[tuple[str, str]]) -> str:
    s = text
    for pat, rep in pairs:
        s = re.sub(pat, rep, s, flags=re.IGNORECASE)
    return s


def transform_description(s: str) -> str:
    if pd.isna(s) or str(s).strip() == "" or str(s) == "nan":
        return ""
    t = str(s).strip()
    t = _sub_chain(t, ORDERED_REPLACEMENTS)
    t = _sub_chain(t, DEV_PHRASES)
    t = _sub_chain(t, POINT_EXTRA)
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def transform_logic(s: str, _: str) -> str:
    if pd.isna(s) or str(s).strip() == "" or str(s) == "nan":
        return ""
    t = str(s).strip()
    t = _sub_chain(t, ORDERED_REPLACEMENTS)
    t = _sub_chain(t, DEV_PHRASES)
    # 箭头留白，便于朗读
    t = t.replace("→", " → ")
    t = re.sub(r"\s*→\s*", " → ", t)
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def transform_label(s: str) -> str:
    if pd.isna(s) or str(s).strip() == "" or str(s) == "nan":
        return ""
    t = str(s).strip()
    t = _sub_chain(t, ORDERED_REPLACEMENTS)
    t = _sub_chain(t, POINT_EXTRA)
    return t.strip()


def transform_remark(s: str) -> str:
    """备注列：源表无则空；若有则通俗化。"""
    if pd.isna(s) or str(s).strip() == "" or str(s) == "nan":
        return ""
    t = str(s).strip()
    t = _sub_chain(t, ORDERED_REPLACEMENTS)
    t = _sub_chain(t, DEV_PHRASES)
    # 长英文栈 → 简短
    t = re.sub(
        r"依赖\s*OpenAI[^；。]*",
        "依赖大模型云与向量库，按项目实际选型",
        t,
        flags=re.IGNORECASE,
    )
    t = re.sub(
        r"Email\([^)]+\)\s*\+\s*[^；]+兜底",
        "邮件与手机推送作补充提醒",
        t,
    )
    t = re.sub(
        r"界面与推送支持[^；]+",
        "界面与消息支持俄/乌/英/中，默认俄语",
        t,
    )
    t = re.sub(
        r"全局时区[^；]+",
        "全站时间按塔什干时区展示与统计",
        t,
    )
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def style_workbook(path: Path, nrows: int) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(ws[1]))
    ws.auto_filter.ref = f"A1:{last_col}{nrows + 1}"

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    fonts = Font(name="Microsoft YaHei", size=10)
    alt_a = PatternFill("solid", fgColor="FFFFFF")
    alt_b = PatternFill("solid", fgColor="F2F2F2")

    widths = {"A": 10, "B": 22, "C": 26, "D": 46, "E": 68, "F": 36}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ports: list[str] = []
    for row in range(2, nrows + 2):
        ports.append(str(ws.cell(row=row, column=1).value or ""))

    toggle = 0
    prev = None
    for idx, p in enumerate(ports):
        r = idx + 2
        if p != prev:
            toggle = 1 - toggle
            prev = p
        fill = alt_a if toggle == 0 else alt_b
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = fonts
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = fill

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    wb.save(path)


def main() -> None:
    df = pd.read_excel(SRC)
    assert list(df.columns) == ["端口", "板块", "功能点", "详细功能描述", "功能逻辑"], list(df.columns)

    df = df.copy()
    df["端口"] = df["端口"].ffill()
    df["板块"] = df["板块"].ffill()

    out = pd.DataFrame()
    out["端口"] = df["端口"].map(transform_label)
    out["板块"] = df["板块"].map(transform_label)
    out["功能点"] = df["功能点"].map(transform_label)
    out["详细功能描述"] = df["详细功能描述"].map(transform_description)
    out["功能逻辑"] = [
        transform_logic(r["功能逻辑"], "") for _, r in df.iterrows()
    ]
    out["备注"] = pd.Series([""] * len(out), dtype=object)

    out.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="通俗版需求", na_rep="")

    style_workbook(OUT_XLSX, len(out))

    OUT_TXT.write_text(
        "\n".join(
            [
                "《需求功能清单·通俗版》使用说明",
                "",
                "1. 本文件在「主清单 xlsx」基础上改写用语，行数、排序与主清单一致。",
                "2. 适合商务、教务客户阅读；技术同事仍可从「功能逻辑」把握实现要点。",
                "3. 已尽量减少英文与技术缩写；保留 Telegram、Payme、Click 等客户常见品牌名。",
                "4. 货币统一写作「乌兹别克苏姆」；时间统一为「塔什干时间（乌兹别克）」。",
                "5. 「备注」列：主 xlsx 原表无此列，本版暂留空，可在评审时由双方补充优先级或依赖。",
                "6. 端口、板块列已按合并单元格习惯向前填充，便于筛选与阅读。",
                "7. CSV 为 UTF-8（带 BOM），可直接用 Excel 打开。",
                "",
                "（说明由生成脚本附带，可随项目修订。）",
            ]
        ),
        encoding="utf-8",
    )

    print("rows", len(out))
    print("out", OUT_XLSX.resolve())
    print("csv", OUT_CSV.resolve())


if __name__ == "__main__":
    main()
